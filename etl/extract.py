#!/usr/bin/env python3
"""
ETL: normalize 'ใบเบิกสินค้าเครือ Central.xlsx' -> clean JSON seed files.

Run:  python3 etl/extract.py
Out:  seed/*.json  (products, branches, purchase_orders, po_items,
                    shipment_items, return_items, sales, daily_customers,
                    product_costs, cash_entries, stock_snapshot, meta)

Design notes
------------
- The workbook mixes marker rows ('***'), Excel errors ('#REF!', '#N/A'),
  Buddhist-era dates (year > 2500) and blank spacer rows. We defensively
  clean all of these.
- Output is plain JSON with ISO date strings so a Node seed script can load
  it straight into Postgres/PGlite/Supabase without re-parsing Excel.
"""
import json, os, datetime, re
from decimal import Decimal
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "source.xlsx")
OUT = os.path.abspath(os.path.join(HERE, "..", "seed"))
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)

ERR = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "***", "", "-"}


def clean(v):
    """Normalize a raw cell value."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s in ERR:
            return None
        return s
    return v


def as_date(v):
    """Return ISO date string, correcting Buddhist-era years (e.g. 2568 -> 2025)."""
    if isinstance(v, datetime.datetime):
        y = v.year
        if y > 2500:
            try:
                v = v.replace(year=y - 543)
            except ValueError:
                v = v.replace(year=y - 543, day=28)
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return None


def as_time(v):
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, datetime.datetime):
        return v.time().strftime("%H:%M:%S")
    return None


def num(v):
    """Best-effort numeric -> float, else None."""
    v = clean(v)
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def barcode(v):
    """Barcodes come as floats (8857128011188.0) or strings; keep as 13-digit str."""
    v = clean(v)
    if v is None:
        return None
    if isinstance(v, float):
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    return s if s else None


def rows(sheet, start):
    ws = wb[sheet]
    for r in range(start, ws.max_row + 1):
        yield r, ws


# ---------------------------------------------------------------- products
def extract_products():
    ws = wb["Main Data Base"]
    out, seen = [], set()
    for r in range(2, ws.max_row + 1):
        bc = barcode(ws.cell(r, 1).value)          # A Barcode
        scent = clean(ws.cell(r, 2).value)         # B Scent
        if not bc or not scent or bc in seen:
            continue
        seen.add(bc)
        out.append({
            "barcode": bc,
            "scent": scent,
            "grade": clean(ws.cell(r, 3).value),   # C Grand
            "size": clean(ws.cell(r, 4).value),    # D Size
            "sku": clean(ws.cell(r, 5).value),     # E Code
            "brand": clean(ws.cell(r, 6).value) or "Lab Parfumo",  # F
            "price": num(ws.cell(r, 7).value),     # G ราคา
            "description": clean(ws.cell(r, 8).value),  # H
        })
    return out


# ---------------------------------------------------------------- branches
def extract_branches():
    """Branches live in Main Data Base cols L..T, plus any seen in PO List."""
    ws = wb["Main Data Base"]
    out, seen = [], set()
    for r in range(2, ws.max_row + 1):
        code = clean(ws.cell(r, 12).value)         # L Branch Code
        if not code or code in seen:
            continue
        seen.add(code)
        out.append({
            "branch_code": code,
            "store_code": clean(ws.cell(r, 13).value),   # M
            "store_no": clean(ws.cell(r, 14).value),     # N Branch (สาขาที่)
            "tel": clean(ws.cell(r, 15).value),          # O
            "receiver": clean(ws.cell(r, 16).value),     # P
            "email_store": clean(ws.cell(r, 17).value),  # Q
            "email_admin": clean(ws.cell(r, 18).value),  # R
            "address": clean(ws.cell(r, 19).value),      # S
        })
    return out


# ------------------------------------------------------ purchase orders / items
def extract_po():
    """PO List Lab Parfumo -> header per (po_number,version) + line items."""
    ws = wb["PO List Lab Parfumo"]
    headers, items = {}, []
    for r in range(4, ws.max_row + 1):
        po = clean(ws.cell(r, 3).value)            # C PO Order Number
        if not po:
            continue
        version = clean(ws.cell(r, 4).value)       # D version
        date = as_date(ws.cell(r, 2).value)        # B Date
        branch = clean(ws.cell(r, 5).value)        # E Branch
        key = (po, version or "")
        if key not in headers:
            headers[key] = {
                "po_number": po,
                "version": version,
                "order_date": date,
                "branch_label": branch,
                "store_no": clean(ws.cell(r, 6).value),      # F Branches
                "delivery_number": clean(ws.cell(r, 11).value),  # K
                "phone": clean(ws.cell(r, 12).value),        # L
                "shipping_name": clean(ws.cell(r, 13).value),  # M
                "address": clean(ws.cell(r, 14).value),      # N
                "remark": clean(ws.cell(r, 15).value),       # O
            }
        items.append({
            "po_number": po,
            "version": version,
            "line_no": num(ws.cell(r, 1).value),   # A Order
            "barcode": barcode(ws.cell(r, 7).value),  # G
            "scent": clean(ws.cell(r, 8).value),   # H Scent
            "size": clean(ws.cell(r, 9).value),    # I Size
            "qty": num(ws.cell(r, 10).value) or 0,  # J จำนวน
        })
    return list(headers.values()), items


# --------------------------------------------------- shipment / return items
def extract_shipments():
    ws = wb["รายการส่งสินค้า"]
    out = []
    for r in range(2, ws.max_row + 1):
        po = clean(ws.cell(r, 3).value)            # C Order Number
        name = clean(ws.cell(r, 5).value)          # E Name
        if not po and not name:
            continue
        out.append({
            "line_no": num(ws.cell(r, 1).value),   # A
            "ship_date": as_date(ws.cell(r, 2).value),  # B
            "po_number": po,
            "sku": clean(ws.cell(r, 4).value),     # D SKU
            "name": name,
            "serial": barcode(ws.cell(r, 6).value),  # F Serial
            "grade": clean(ws.cell(r, 7).value),   # G
            "size": clean(ws.cell(r, 8).value),    # H
            "branch_label": clean(ws.cell(r, 9).value),  # I
            "receive_status": clean(ws.cell(r, 11).value),  # K
        })
    return out


def extract_returns():
    ws = wb["รายการคืนสินค้า"]
    out = []
    for r in range(2, ws.max_row + 1):
        po = clean(ws.cell(r, 4).value)            # D Order Number
        name = clean(ws.cell(r, 6).value)          # F Name
        if not po and not name:
            continue
        out.append({
            "line_no": num(ws.cell(r, 1).value),   # A
            "return_date": as_date(ws.cell(r, 2).value),  # B Shipping
            "po_number": po,
            "sku": clean(ws.cell(r, 5).value),     # E SKU
            "name": name,
            "serial": barcode(ws.cell(r, 7).value),  # G
            "grade": clean(ws.cell(r, 8).value),   # H
            "size": clean(ws.cell(r, 9).value),    # I
            "branch_label": clean(ws.cell(r, 10).value),  # J
            "receive_status": clean(ws.cell(r, 12).value),  # L
        })
    return out


# ---------------------------------------------------------------- sales
def extract_sales():
    out = []
    for sheet, channel in (("Sell CTW", "CTW"), ("Sell EVENT SCS", "EVENT_SCS")):
        ws = wb[sheet]
        for r in range(3, ws.max_row + 1):
            item = clean(ws.cell(r, 6).value)      # F รายการ
            bc = barcode(ws.cell(r, 7).value)      # G Barcode
            if not item and not bc:
                continue
            nation = clean(ws.cell(r, 20).value)   # T Nation
            if isinstance(nation, str):
                nation = nation.capitalize()
            out.append({
                "source": channel,
                "month": clean(ws.cell(r, 1).value),   # A Month
                "sale_date": as_date(ws.cell(r, 2).value),  # B วันที่
                "ba": clean(ws.cell(r, 3).value),      # C BA
                "order_no": clean(ws.cell(r, 4).value),  # D
                "receipt_no": clean(ws.cell(r, 5).value),  # E
                "item": item,
                "barcode": bc,
                "grade": clean(ws.cell(r, 8).value),   # H Grand
                "size": clean(ws.cell(r, 9).value),    # I Size
                "qty": num(ws.cell(r, 10).value) or 0,  # J
                "unit_price": num(ws.cell(r, 11).value),  # K ราคา
                "discount": num(ws.cell(r, 12).value) or 0,  # L
                "total": num(ws.cell(r, 13).value),    # M รวม
                "paid": num(ws.cell(r, 14).value),     # N ยอดชำระ
                "payment_channel": clean(ws.cell(r, 15).value),  # O
                "sale_time": as_time(ws.cell(r, 17).value),  # Q Time
                "note": clean(ws.cell(r, 19).value),   # S หมายเหตุ
                "nation": nation,
            })
    return out


# ------------------------------------------------------- daily customers
def extract_customers():
    ws = wb["Customer"]
    out = []
    for r in range(2, ws.max_row + 1):
        date = as_date(ws.cell(r, 2).value)        # B Date
        if not date:
            continue
        out.append({
            "month": clean(ws.cell(r, 1).value),   # A
            "cust_date": date,
            "ba": clean(ws.cell(r, 3).value),      # C BA
            "customers": int(num(ws.cell(r, 4).value) or 0),  # D
            "sell_amount": num(ws.cell(r, 5).value) or 0,     # E
            "thai": num(ws.cell(r, 6).value),      # F
            "thai_sales": num(ws.cell(r, 7).value),  # G TBGS
            "foreign": num(ws.cell(r, 8).value),   # H
            "foreign_sales": num(ws.cell(r, 9).value),  # I FBGS
        })
    return out


# ---------------------------------------------------------- product costs
def extract_costs():
    ws = wb["ข้อมูลการขาย"]
    out = []
    for r in range(6, ws.max_row + 1):
        scent = clean(ws.cell(r, 3).value)         # C SCENT
        if not scent:
            continue
        out.append({
            "cost_date": as_date(ws.cell(r, 2).value),  # B
            "scent": scent,
            "size": clean(ws.cell(r, 4).value),    # D
            "barcode": barcode(ws.cell(r, 5).value),  # E REFERENCE
            "unit_cost": num(ws.cell(r, 6).value),  # F
            "qty": num(ws.cell(r, 7).value) or 0,  # G
            "total_cost": num(ws.cell(r, 8).value),  # H
        })
    return out


# ------------------------------------------------------------------ cash
def extract_cash():
    ws = wb["Cash"]
    out = []
    for r in range(2, ws.max_row + 1):
        date = as_date(ws.cell(r, 1).value)        # A
        amt = num(ws.cell(r, 3).value)             # C
        if not date and amt is None:
            continue
        out.append({
            "cash_date": date,
            "description": clean(ws.cell(r, 2).value),  # B
            "amount": amt,
            "type": clean(ws.cell(r, 4).value),    # D
        })
    return out


# ---------------------------------------------------- stock snapshot (ref)
def extract_stock():
    ws = wb["สต๊อกสินค้าคงเหลือ"]
    out = []
    for r in range(6, ws.max_row + 1):
        scent = clean(ws.cell(r, 3).value)         # C SCENT
        if not scent:
            continue
        out.append({
            "scent": scent,
            "size": clean(ws.cell(r, 4).value),    # D
            "barcode": barcode(ws.cell(r, 5).value),  # E
            "shipped": num(ws.cell(r, 6).value) or 0,   # F รายการส่งไป
            "sold": num(ws.cell(r, 7).value) or 0,      # G รายการขาย
            "remaining": num(ws.cell(r, 8).value) or 0,  # H คงเหลือ
        })
    return out


def write(name, data):
    path = os.path.join(OUT, name + ".json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f"  {name:20s} {len(data):5d} rows -> seed/{name}.json")


def main():
    print("Extracting ->", OUT)
    products = extract_products()
    branches = extract_branches()
    po_headers, po_items = extract_po()
    write("products", products)
    write("branches", branches)
    write("purchase_orders", po_headers)
    write("po_items", po_items)
    write("shipment_items", extract_shipments())
    write("return_items", extract_returns())
    write("sales", extract_sales())
    write("daily_customers", extract_customers())
    write("product_costs", extract_costs())
    write("cash_entries", extract_cash())
    write("stock_snapshot", extract_stock())
    write("meta", {"generated_from": os.path.basename(SRC),
                   "products": len(products), "branches": len(branches),
                   "purchase_orders": len(po_headers)})
    print("Done.")


if __name__ == "__main__":
    main()
